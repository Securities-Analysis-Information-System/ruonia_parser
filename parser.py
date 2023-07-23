import urllib.request
import openpyxl
import warnings
from datetime import datetime, timedelta
from calendar import monthrange
import pandas as pd
import sys
import re
import os, glob

from visualize_ruonia import visualize_ruonia

DATE_TO_OPTION = 'date_to'
DATE_FROM_OPTION = 'date_from'
MODE_OPTION = 'mode'
BONUS_OPTION = 'bonus'
IS_NEED_VISUALIZE_FLAG = 'visualize'
IS_NEED_DELETE_RUONIA_PARSER_XLSX_FILES_FLAG = 'delete_parser_xlsx_files_in_run_directory'

DATE_FORMAT = '%Y-%m-%d'
ENCODED_DOT = '%2F'

optional_flags = [IS_NEED_VISUALIZE_FLAG, IS_NEED_DELETE_RUONIA_PARSER_XLSX_FILES_FLAG]
optional_args = [BONUS_OPTION, DATE_FROM_OPTION]
required_args = [MODE_OPTION, DATE_TO_OPTION]
possible_values = {
    MODE_OPTION: ['half', 'quadro'],
    DATE_TO_OPTION: '\d{4}-\d{2}-\d{2}',
    DATE_FROM_OPTION: '\d{4}-\d{2}-\d{2}',
    BONUS_OPTION: '\d',
}
founded_optional_flags = {
    IS_NEED_VISUALIZE_FLAG: '0',
    IS_NEED_DELETE_RUONIA_PARSER_XLSX_FILES_FLAG: '0'
}
founded_optional_args = {}
founded_args = {}
errors = []
for argument in sys.argv:
    if argument[:2] == '--' and '=' in argument:
        option = argument.split('=')
        name = option[0].split('--')[1]
        value = option[1]

        if name in required_args and name not in founded_args:
            if name in possible_values:
                if type(possible_values[name]) is list:
                    if value in possible_values[name]:
                        founded_args[name] = value
                    else:
                        errors.append(
                            f'Not valid value "{value}" for option "{name}", possible values: {possible_values[name]}')

                if type(possible_values[name]) is not list:
                    if re.match(possible_values[name], value):
                        founded_args[name] = value
                    else:
                        errors.append(
                            f'Not valid value "{value}" for option "{name}", possible format: {possible_values[name]}')
        if name in optional_args:
            if name in possible_values:
                if type(possible_values[name]) is list:
                    print("HERE1")
                    if value in possible_values[name]:
                        founded_optional_args[name] = value
                    else:
                        errors.append(
                            f'Not valid value "{value}" for option "{name}", possible values: {possible_values[name]}')

                if type(possible_values[name]) is not list:
                    if re.match(possible_values[name], value):
                        founded_optional_args[name] = value
                    else:
                        errors.append(
                            f'Not valid value "{value}" for option "{name}", possible format: {possible_values[name]}')
    elif argument[:2] == '--' and argument[2:] in optional_flags:
        name = argument[2:]

        if name in optional_flags:
            founded_optional_flags[name] = '1'
    else:
        continue

if len(errors) > 0:
    print('Errors with options:')
    for idx, error in enumerate(errors, 1):
        print(f'{idx}) {error}')
    exit(1)

if len(founded_args) != len(required_args):
    notExistsOptions = list(set(required_args) - set(founded_args))
    print(f'This options not specified, but required: {notExistsOptions}')
    exit(2)

date_to_option_formatted = datetime.strptime(founded_args[DATE_TO_OPTION], DATE_FORMAT)
if date_to_option_formatted > datetime.now():
    print(
        f'"date_to" "{date_to_option_formatted.strftime("%Y-%m-%d")}" is greater than current day "{datetime.now().strftime("%Y-%m-%d")}". Please specify "date_to" lower or equal current day.')
    exit(1)

date_from_option_formatted = None
if DATE_FROM_OPTION in founded_optional_args:
    date_from_option_formatted = datetime.strptime(founded_optional_args[DATE_FROM_OPTION], DATE_FORMAT)
    if date_from_option_formatted > date_to_option_formatted:
        print(
            f'"date_from" "{date_from_option_formatted.strftime("%Y-%m-%d")}" is greater than "date_to" "{date_to_option_formatted.strftime("%Y-%m-%d")}". Please specify "date_from" lower or equal "date_to".')
        exit(1)

print(founded_args, founded_optional_args)
founded_args = dict(
    list(founded_args.items()) + list(founded_optional_args.items()) + list(founded_optional_flags.items()))

is_need_visulize = founded_args[IS_NEED_VISUALIZE_FLAG] == '1'
is_need_delete_old_xlsx_files = founded_args[IS_NEED_DELETE_RUONIA_PARSER_XLSX_FILES_FLAG] == '1'
end_date = date_to_option_formatted
start_date = date_from_option_formatted
mode = founded_args[MODE_OPTION]
bonus: float = 0
if BONUS_OPTION in founded_args:
    bonus = float(founded_args[BONUS_OPTION])

if is_need_delete_old_xlsx_files:
    for oldXlsx in glob.glob('ruonia-parser_*_*.xlsx'):
        os.remove(oldXlsx)

delta = 91
if mode == 'half':
    end_date = end_date - timedelta(days=182)
    delta = 182
if mode == 'quadro':
    delta = 91

if start_date is None:
    start_date = end_date - timedelta(days=delta)

startYear = start_date.year.real
months = []
previousMonth = None
startDay = start_date.day.real
endDay = end_date.day.real
for date in pd.date_range(start_date, end_date).tolist():
    if previousMonth != date.month.real:
        months.append(date.month.real)
        previousMonth = date.month.real

firstMonth = months[0]
lastMonth = months[len(months) - 1]

year = startYear
count = 0
total_ruonia = 0
countBad = 0
previousMonth = 11
already = False

data = []
avg_data = []
for month in months:
    monthStr = month
    if month < 10:
        monthStr = f'0{month}'

    if previousMonth == 12:
        year = f'{int(year) + 1}'
    previousMonth = month
    month_range = monthrange(int(year), month)

    dayPeriodBegin = '01'
    if month == firstMonth and already is False:
        dayPeriodBegin = startDay
        if dayPeriodBegin < 10:
            dayPeriodBegin = f'0{dayPeriodBegin}'
        already = True
    dayPeriodEnd = month_range[1]
    if month == lastMonth and already is True:
        dayPeriodEnd = endDay
        if dayPeriodEnd < 10:
            dayPeriodEnd = f'0{dayPeriodEnd}'

    dateFromEDot = f'{monthStr}{ENCODED_DOT}{dayPeriodBegin}{ENCODED_DOT}{year}'
    dateToEDot = f'{monthStr}{ENCODED_DOT}{dayPeriodEnd}{ENCODED_DOT}{year}'

    url = f'https://cbr.ru/Queries/UniDbQuery/DownloadExcel/14315?Posted=True&FromDate={dateFromEDot}&ToDate={dateToEDot}'
    filename = f'ruonia-parser_{dayPeriodBegin}{monthStr}{year}_{dayPeriodEnd}{monthStr}{year}.xlsx'

    urllib.request.urlretrieve(
        url,
        filename
    )

    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        book = openpyxl.load_workbook(filename)
        sheet = book.active

    date_from = datetime.strptime(f'{year}-{monthStr}-{dayPeriodBegin}', '%Y-%m-%d')
    date_to = datetime.strptime(f'{year}-{monthStr}-{dayPeriodEnd}', '%Y-%m-%d')
    date_list = pd.date_range(date_from, date_to).strftime('%Y-%m-%d').tolist()
    date_list.reverse()

    dateColumn = 1
    ruoniaColumn = 2
    row = 2
    tupler = []
    tupler_avg = []
    for expected_date in date_list:
        ruonia: float = sheet.cell(row=row, column=ruoniaColumn).value
        date: datetime = sheet.cell(row=row, column=dateColumn).value
        formattedDate = date
        if date != None:
            formattedDate = date.strftime('%Y-%m-%d')

        if expected_date != formattedDate:
            print(f'Bad ass. {expected_date} - {formattedDate}')
        else:
            count = count + 1

            ruonia = ruonia
            total_ruonia = total_ruonia + ruonia

            print(f'{formattedDate}: {ruonia} / {expected_date}')

            # if count % 2 == 0:
            tupler.append((formattedDate, ruonia))
            tupler_avg.append((formattedDate, 0))
            # else:
            #     data.append(('', ruonia))

            row = row + 1

    tupler.reverse()
    tupler_avg.reverse()
    for item in tupler:
        data.append(item)

    for item in tupler_avg:
        avg_data.append(item)

avg_ruonia = total_ruonia / count + bonus

print()
print(total_ruonia, count, avg_ruonia)

new_avg_data = []
for item in avg_data:
    new_item = (item[0], avg_ruonia)
    new_avg_data.append(new_item)

avg_data = new_avg_data

if is_need_visulize:
    visualize_ruonia([data, avg_data], count)
